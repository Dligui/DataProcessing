# Form implementation generated from reading ui file 'CopyData.ui'
#
# Created by: PyQt6 UI code generator 6.4.0
#
# WARNING: Any manual changes made to this file will be lost when pyuic6 is
# run again.  Do not edit this file unless you know what you are doing.

import win32com.client
from PyQt6.QtCore import *
from PyQt6.QtGui import *
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QDialog,QApplication,QFileDialog,QMainWindow,QComboBox
from sorted_tab import sorted_tab
from my_head import my_head
import sys
from Paste_table_vf import Paste_table_vf
from PyQt6.QtGui import QIcon
from openpyxl import  load_workbook 
from Nb_tests import Nb_tests

class Ui_DataProcessing(object):

    # ExcelApp = win32com.client.GetActiveObject("Excel.Application")
    # ExcelApp.Visible = True

    def __init__(self):
        super(Ui_DataProcessing,self).__init__()

    def Browsefiles(self):
        fname=QFileDialog.getOpenFileName(None,'open file','.')
        self.Filename.setText(fname[0])
        path=self.Filename.text()
        # wb=load_workbook(path)
        # sheets_names=wb.sheetnames
        # for i in sheets_names:
        #     self.sheets.addItem(i)
        # sheet=str(self.sheets.currentText())
        # sheet="'"+sheet+"'" 
        # print(sheet)
        #data=my_head(path,sheet)

        # head1=list(Nb_tests(path)[1])
        # print(head1)
        # self.comboBox1.addItems(head1[0])

        sheets=Nb_tests(path)[0]
        self.comboBox.addItems(sheets)
                
    def add_item(self):
        indx=[]
        B=self.checkBox.isChecked()
        ind=self.comboBox.currentIndex()
        self.listWidget.addItem(str(ind)+" : "+str(self.comboBox.currentText()) + " will be sorted in ascending order? "+ str(B))
        indx.extend(str(ind))
        indx = [int(i) for i in indx] 
        return indx
  
    def delete_item(self):
        list_item=self.listWidget.selectedItems()
        if not list_item: return
        for item in list_item:
            item_index=self.listWidget.row(item)
            self.listWidget.takeItem(item_index)
        return item_index

    def setupUi(self, DataProcessing):
        DataProcessing.setObjectName("DataProcessing")
        DataProcessing.resize(922, 412)
        self.centralwidget = QtWidgets.QWidget(DataProcessing)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 20, 131, 61))
        self.label.setObjectName("label")

        self.CopyData = QtWidgets.QPushButton(self.centralwidget)
        self.CopyData.setGeometry(QtCore.QRect(640, 90, 151, 51))
        font = QtGui.QFont()
        font.setBold(True)
        font.setItalic(False)
        font.setUnderline(False)
        font.setWeight(75)

        self.CopyData.setAutoFillBackground(True)
        self.CopyData.setObjectName("CopyData")
        self.CopyData.setStyleSheet("background-color:green;\n"
                                      "color: white;\n"
                                      "border-style: outset;\n"
                                      "border-width:2px;\n"
                                      "border-radius:10px;\n"
                                      "border-color:black;\n"
                                      "font:bold 14px;\n"
                                      "padding :6px;\n"
                                      "min-width:10px;\n"
                                      "\n"
                                      "\n"
                                      "")
        self.CopyData.clicked.connect(self.paste_vf)
       
        self.Browse = QtWidgets.QPushButton(self.centralwidget)
        self.Browse.setGeometry(QtCore.QRect(640, 30, 151, 51))
        self.Browse.setAutoFillBackground(True)
        self.Browse.setObjectName("Browse")
        self.Browse.setStyleSheet("background-color:red;\n"
                                      "color: white;\n"
                                      "border-style: outset;\n"
                                      "border-width:2px;\n"
                                      "border-radius:10px;\n"
                                      "border-color:black;\n"
                                      "font:bold 14px;\n"
                                      "padding :6px;\n"
                                      "min-width:10px;\n"
                                      "\n"
                                      "\n"
                                      "")
        self.Browse.clicked.connect(self.Browsefiles)

        self.button = QtWidgets.QPushButton(self.centralwidget)
        self.button.setGeometry(QtCore.QRect(640, 150, 151, 51))
        self.button.setAutoFillBackground(True)
        self.button.setObjectName("button")
        self.button.setStyleSheet("background-color:yellow;\n"
                                      "color: black;\n"
                                      "border-style: outset;\n"
                                      "border-width:2px;\n"
                                      "border-radius:10px;\n"
                                      "border-color:black;\n"
                                      "font:bold 14px;\n"
                                      "padding :6px;\n"
                                      "min-width:10px;\n"
                                      "\n"
                                      "\n"
                                      "")

        self.Filename = QtWidgets.QLineEdit(self.centralwidget)
        self.Filename.setGeometry(QtCore.QRect(70, 30, 551, 51))
        self.Filename.setObjectName("Filename")

        self.checkBox = QtWidgets.QCheckBox(self.centralwidget)
        self.checkBox.setGeometry(QtCore.QRect(290, 160, 121, 41))
        self.checkBox.setObjectName("checkBox")

        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(70, 160, 181, 31))
        self.comboBox.setObjectName("comboBox")

        self.comboBox1 = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox1.setGeometry(QtCore.QRect(70, 200, 181, 31))
        self.comboBox1.setObjectName("comboBox1")

        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(70, 90, 551, 51))
        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.MultiSelection)
        self.listWidget.setObjectName("listWidget")
        
        self.buttonBox = QtWidgets.QDialogButtonBox(self.centralwidget)
        self.buttonBox.setGeometry(QtCore.QRect(430, 160, 193, 28))
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.StandardButton.Cancel|QtWidgets.QDialogButtonBox.StandardButton.Ok)
        self.buttonBox.setObjectName("buttonBox")
        self.buttonBox.accepted.connect(self.add_item)
        #self.buttonBox.accepted.connect(self.indx)
        self.buttonBox.rejected.connect(self.delete_item)
        # self.buttonBox.rejected.connect(self.indx)

        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 80, 131, 61))
        self.label_2.setObjectName("label_2")

        self.menubar = QtWidgets.QMenuBar(DataProcessing)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 922, 26))
        self.menubar.setObjectName("menubar")

        self.statusbar = QtWidgets.QStatusBar(DataProcessing)
        self.statusbar.setObjectName("statusbar")

        self.retranslateUi(DataProcessing)
        QtCore.QMetaObject.connectSlotsByName(DataProcessing)

    def retranslateUi(self, DataProcessing):
        _translate = QtCore.QCoreApplication.translate
        DataProcessing.setWindowTitle(_translate("DataProcessing", "MainWindow"))
        self.label.setText(_translate("DataProcessing", "Path : "))
        self.CopyData.setText(_translate("DataProcessing", "CopyData"))
        self.Browse.setText(_translate("DataProcessing", "Browse"))
        self.checkBox.setText(_translate("DataProcessing", "Ascending"))
        self.label_2.setText(_translate("DataProcessing", "Colums : "))
        self.button.setText(_translate("DataProcessing", "list"))

    def paste_vf(self):
        path=self.Filename.text()
        print(path)
        #L=self.add_item() # delete this combobox
        B=self.checkBox.isChecked()
        DataRoom=self.comboBox.currentText()
        Paste_table_vf(path,B,[1],DataRoom)

if __name__== "__main__":
    import sys
    app=QtWidgets.QApplication(sys.argv)
    form=QtWidgets.QWidget()
    ui=Ui_DataProcessing()
    ui.setupUi(form)
    form.show()
    sys.exit(app.exec())