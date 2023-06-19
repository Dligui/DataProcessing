from openpyxl.workbook import Workbook
from openpyxl import  load_workbook 
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange1 import pasteRange1
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate
from Tab_Sheets import Tab_sheets
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from copy_head import  copy_head
from paste_table import paste_table
from sorted_tab import sorted_tab
from my_head import my_head
import itertools

def Nb_tests(path):
    wb=load_workbook(path) # ,read_only=False, keep_vba=True)
    sheets_name=wb.sheetnames
    list_tests=[]
    list_sheet=[]
    autors=[]
    for j in range(0,len(sheets_name)):
        ws = wb[sheets_name[j]]
        maxr=ws.max_row
        maxc=ws.max_column
        info=infosheet(maxr,maxc,ws)
        endCol=info[5]
        startrowcol=info[1]
        for i in range(0,info[0]):
            head=copyRange(startrowcol[i][1],startrowcol[i][0]+1,endCol[i],info[3],ws)    
            list_tests.append(head)
            autor=copyRange(startrowcol[i][1],startrowcol[i][0]-1,endCol[i],info[3]-2,ws)
            autors.append(autor[0][0])
            list_sheet.append(sheets_name[j])
        list_sheet=list(set(tuple(list_sheet))) 
    list_tests=list(list_tests for list_tests,_ in itertools.groupby(list_tests))
    return list_sheet,list_tests,autors   
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# print(Nb_tests(path)[0])
# print("*******************************************************************************")
# print(Nb_tests(path)[1])
# print("*******************************************************************************")
# print(Nb_tests(path)[2])

