from openpyxl.workbook import Workbook
from openpyxl import  load_workbook 
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
import xlwt
import win32com.client 
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange1 import pasteRange1
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate
from Tab_Sheets import Tab_sheets
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from copy_head import  copy_head
from paste_table import paste_table
from sorted_tab import sorted_tab
from operator import is_not
from functools import partial
import xlwings as xw

def Paste_table_vf(path,B,L,DataRoom):
    wb=load_workbook(path)
    #wb=xw.books.active
    tabs=sorted_tab(path,B,L,DataRoom)
    ws2=wb['Calculus']  
    #ws2=wb.sheets.active  
    paste_table(path,ws2,tabs[1])
    wb.save(path)
    print('Finished...')

path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx' # 
Paste_table_vf(path,True,[1],'DataRoom')

#tabs=copy_head(path,DataRoom) 
#tabs.dropna(axis=0,inplace=True) # delete NaN
#tabs=list(filter(partial(is_not, None), tabs))