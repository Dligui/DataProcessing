from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
def Tab_sheets(path):
    wb=load_workbook(path) # ,read_only=False, keep_vba=True)
    sheets=wb.sheetnames
    ws = wb.active
    infos=[]
    i=0
    datasheet=[]
    for sheet in sheets:
        ws = wb.get_sheet_by_name(name = sheet )
        maxr=ws.max_row
        maxc=ws.max_column
        infos.append(infosheet(maxr,maxc,ws))
        if infos[i][0]!=0 :
            datasheet.append(sheet)
            datasheet.append(infos[i])
        i=i+1
    return datasheet
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# print(Tab_sheets(path))
