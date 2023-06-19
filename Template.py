from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet

#*****************************Create 6 worksheets for our template workbook***************************
wb=load_workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx')
ws = wb.get_sheet_by_name(name = 'DataRoom1')
ws1 = wb.get_sheet_by_name(name = 'calc')
ws2= wb.get_sheet_by_name(name = 'temp')

# print (wb.sheetnames)
# print(wb.worksheets[0]) #Summary
# g_sheet=wb.sheetnames
# print(g_sheet)
#ws=wb.active                                          
#ws=wb.worksheets[7] # DataRoom
maxr=ws.max_row
maxc=ws.max_column
cmp=0
#df1=pd.DataFrame()
head_index=[]
tab_index=[]
data=[]                                                                                                   
tab=[]
#*********************copy data from dataroom1 to calc1**********************
#copyall(maxr,maxc,ws,ws1)

#*****************copy data from dataroom1 to temp****************************
# for r in range (1, maxr + 1):
#     for c in range (1, maxc + 1):
#         if (ws.cell(row=r,column=c).value=='Data'):
#             print('**************here*******************')
#             for r in range(1,maxr+1):
#                 for j in 
#                 print(ws.cell(row=r,column=c).value)
#                           
#**************get_table**************************************************


whs=getTable(t,h,maxr,cmp,ws)
print(whs)
rang=copyRange(1,5,4,20,ws)
print(rang)
#*************copy all ranges in rang************************************************
n=0  
cmp=3 
rang=[]
while n<cmp:
    a=t[n]
    b=h[n]
    rang.append(copyRange(1,a,b,20,ws))    
    n=n+1

print(rang)
pasteRange(1,1,4,60,ws2,rang)


#pasteRange(1,20,4,35,ws2,rang)
# data=ws.cell(row=i,column=j).value
# df1=df1.concat(df1)
#print(b+1)
# def get_table(b,tab_index):
#     cmp=0
#     for i in range(b+1,tab_index[0][0]):
#         for j in range(1,tab_index[0][0]):
#             cmp=cmp+1
#             tab.append(ws.cell(row=i,column=j).value)           
#     print(tab)

               
# first table**************************
# for i in range(b+1,maxr+1):
#     x=tab_index[0][1]        
#     for j in range(1,x+1):
#         t.append(ws.cell(row=i,column=j).value)

# print(df1)
df1.to_excel('C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx',sheet_name='calc',index=0)
wb.save('C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx')

#print(t)
#***************************************************

#get_table(b,tab_index)
#print(t[0])


# #change xxx with the sheet name that includes the data
# data = pd.read_excel(wb, sheet_name="DataRoom")

# #save it to the 'new_tab' in destfile
# data.to_excel(wb, sheet_name='calculus')



# for [i,j] in tab_index:
#     for r in range(i)
#     





# for c in range (1, maxc + 1):
#     for r in range (1, maxr + 1):
#         if (ws.cell(row=r,column=c).value=='Data'):

#             for r in range (1,maxr+1):
#                 a.append((ws.cell(row=r,column=c).value))
#             df1=pd.DataFrame(a)           
        
#create functions test them and import them here
# df1 = df[df['Data frame']==0]
# df2 = df[df['Data frame']==1]
# df1.to_excel(writer,sheet_name='zero')
# df2.to_excel(writer,sheet_name='one')

# recuperer la liste pour chaque tab 
# nombre de tab via 'Data' : c : done
#min max de dataroom : done 
# min et max de chaq tab 
# developp strong code by testing it via different fils


#*********************************************
# wb=load_workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/temp_py.xlsx')
# ws=wb.active
# wb.copy_worksheet(from_worksheet='DataRoom': 'Calculus')
# c1 = ws.cell(row = 2, column = 1, )
# print(c1)

# x1=pd.ExcelFile('C:/Users/zbptmj/Desktop/newdesk/py_gui/temp_py.xlsx')
# print(x1.sheet_names) # Read sheet names list
# x1.parse(x1.sheet_names[3])
# a=pd.read_excel(io='C:/Users/zbptmj/Desktop/newdesk/py_gui/temp_py.xlsx',index_col=0 ,sheet_name='DataRoom')
# print(a)
# x1.sheet_names[2]=a
#wb.create_sheet(index = 1 , title = "demo sheet2")

# for i in range (1, mr + 1):
#     for j in range (1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row = i, column = j)
  
#         # writing the read value to destination excel file
#         ws2.cell(row = i, column = j).value = c.value