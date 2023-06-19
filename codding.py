from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import time
from tkinter import messagebox
import xlrd
import numpy as np
from openpyxl.styles import Alignment  
import xlwings as xw
from openpyxl.styles import PatternFill

# from FilterColumn import add_filter_column 
#********************Read data brut************** Copy it to DataRoom ws4 ****************************
a=pd.read_excel(io='C:/Users/zbptmj/Desktop/newdesk/py_gui/Data_brut.xlsx')
#b=pd.read_csv('C:/Users/zbptmj/Desktop/newdesk/py_gui/config.ini')
df=pd.DataFrame(a)
# Write DataFrame to Excel file with sheet name 
df.to_excel('C:/Users/zbptmj/Desktop/newdesk/py_gui/py_template.xlsx', sheet_name='DataRoom',index=0)
#*****************************Create 6 worksheets for our template workbook***************************
wb=load_workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/py_template.xlsx')
ws=wb.active
ws.merge_cells('A1:k1')
c=ws.cell(row=1,column=1)
c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
ws1 = wb.create_sheet( "Summary")
ws2 = wb.create_sheet("Report") 
ws3 = wb.create_sheet("Calculus") 
ws4 = wb.create_sheet("Test Configuration")
ws5 = wb.create_sheet("References")
#********************Copy data from DataRoom to Calculus*************************
maxr=ws.max_row
maxc=ws.max_column
#this is work but copy blank cells 
for r in range (3, maxr + 1):
    for c in range (1, maxc + 1):
        ws3.cell(row=r,column=c).value = ws.cell(row=r,column=c).value

# for r in range (3, maxr + 1):
#     for c in range (1, maxc + 1):
#         if ws3.cell(row=r,column=c).value != None : # values_only=True        
#             break            
#         else:        
#             ws3.cell(row=r,column=c).value = ws.cell(row=r,column=c).value

ws3.merge_cells('A1:K1')  
c3=ws3.cell(row=1,column=1)
ws3['A1']= 'onsemi                               Internal Use Only'
c3.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
ws3.auto_filter.ref="A3:k3"

#****************************Summary**********************
ws1.merge_cells('A1:D1')
ws1.merge_cells('A3:D3')
ws1.merge_cells('A4:D4')
ws1.merge_cells('A5:D5')
ws1.merge_cells('A8:D8')
ws1.merge_cells('A9:D9')


c1=ws1.cell(row=4,column=1)
ws1['A4']=ws.cell(row=3,column=10).value # J3
c1.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
ws1['A1']='TLSE Design Validation Lab'
ws1['A3']='TEST   BENCH'
ws1['A5']='DAUGHTER CONFIGURATION : '
ws1['A6']= '=Calculus!$A$25'   # "ws4.cell(row=1,column=3).value "
ws1['A8']='PART :'

c1.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

#ws3.auto_filter.add_filter_column(0,"0",0, '  ')
#Create a data-validation object with list validation
# dv = DataValidation(allow_blank=False)
# ws3.add_data_validation(dv)

# If you want to check for not empty cell use this
# if worksheet.cell(row, col).value != None:

# trg=wb1[ws] 
# src=wb[ws4]
# trg=wb.copy_worksheet(src)
#sheet['A1'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
wb.save('C:/Users/zbptmj/Desktop/newdesk/py_gui/py_template.xlsx')


# for x in range(10):
#     ws.merge_cells(start_row=x, start_column=1, end_row=x, end_column=4)

# pasteRange(1, 5,4, 20, ws2,tab[0])
# pasteRange(1, 21,4, 35, ws2,tab[1])
# pasteRange(1, 36,4, 50, ws2,tab[2])