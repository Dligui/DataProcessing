from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate
from my_head import my_head
import numpy as np
#**************************pyuic6 -x CopyData.ui -o CopyData.py*****************
def copy_head(path,DataRoom):
    tab=[]
    head=my_head(path,DataRoom)
    tab=recuperate(path)
    head.extend(tab)
    return head
#    tab = list(filter(lambda x: str(x) != 'None', tab))
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# wb=load_workbook(path)
# ws = 'DataRoom'
# tab=copy_head(path,ws)
# #tab.remove(np.nan)
# print(tab)
