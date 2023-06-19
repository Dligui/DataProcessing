from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment

#**************************pyuic6 -x CopyData.ui -o CopyData.py*****************
def copy_paste_ranges(path,DataRoom,Calculus):
    wb=load_workbook(path)
    sheets_names=wb.sheetnames
    ws = wb[ DataRoom]
    ws2= wb[Calculus]
    maxr=ws.max_row
    maxc=ws.max_column

    info=infosheet(maxr,maxc,ws)
    startCols=info[4]
    startRows=info[1][0][0]+1
    endCol=info[5]
    endCol[-1]=endCol[-1]-1
    endRow=info[2][0][0]
    totab=info[0]
    tab=[]
    tabsize=endRow-startRows
    head=copyRange(1,5,endCol[0],5,ws)
    for m in range(0,totab-1):
        tab.append(copyRange(startCols[m],startRows+1,endCol[m],endRow,ws)) # 0 -> 8
        pasteRange(1,startRows+tabsize*m+1,endCol[0]+1,endRow+tabsize*m,ws2,tab[m])
    #*******last tab**********************
    tab.append(copyRange(startCols[totab-1],startRows+1,endCol[totab-1],endRow,ws))
    pasteRange(1,startRows+tabsize*(totab-1),endCol[0]+1,endRow+tabsize*(totab-2),ws2,tab[totab-2])
    pasteRange(1,5,endCol[0]+1,5,ws2,head)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1,end_column=ws2.max_column)
    ws2.merge_cells('A3:D3')  
    c1=ws2.cell(row=1,column=1)
    c3=ws2.cell(row=3,column=1)
    ws2['A1']= '  onsemi Internal Use Only'
    ws2['A3']= '   Test Conditions    '
    c1.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    c3.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    FullRange = "A5:" + get_column_letter(ws2.max_column) + str(ws2.max_row)
    ws2.auto_filter.ref=FullRange
    print(tab)
    wb.save(path)
    #return sheets_names

path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
wb=load_workbook(path)
ws1 = 'DataRoom1'
wss='Calculus1'
ws2='DataTemp1'
copy_paste_ranges(path,ws1,wss)
#copy_paste_ranges(path,ws2,wss)
