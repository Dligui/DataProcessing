from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment

#**************************pyuic6 -x CopyData.ui -o CopyData.py*****************
def copy_paste_ranges1(path,DataRoom):
    wb=load_workbook(path) # ,read_only=False, keep_vba=True)
    ws = wb[DataRoom]
    maxr=ws.max_row
    maxc=ws.max_column

    info=infosheet(maxr,maxc,ws)
    startCols=info[4]
    startRows=info[1][0][0]+1
    endCol=info[5]
    endCol[-1]=endCol[-1]-1
    endRow=info[2][0][0]
    totab=info[0]
    tab=[]

    # head=copyRange(1,5,endCol[0],5,ws)    # juste une seule fois
    # tab.extend(head)
    for m in range(0,totab-1): # totab-1
        tab.extend(copyRange(startCols[m],startRows+1,endCol[m],endRow,ws)) # 0 -> 8
    #*******last tab**********************
    tab.extend(copyRange(startCols[totab-1],startRows+1,endCol[totab-1]+1,endRow,ws))

    return tab


#function test
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# wb=load_workbook(path)
# ws = 'DataRoom'
# print(copy_paste_ranges1(path,ws))

#tab=sorted(tabs, key=lambda x: x[4])