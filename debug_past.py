from copyRange import copyRange
from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate

def pasteRange1(startCol, startRow, ws2,rang):
    size=len(rang) # rows
    siz=len(rang[0]) # cols 14
    for i,countRow in zip(range(startRow,size+startRow,1),range(0,size,1)): # size + startRow
        for j, countCol in zip(range(startCol,siz+1,1), range(0,siz+1,1)):  # 1---14 1--13
            ws2.cell(row = i, column = j).value = rang[countRow][countCol]
        
path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template1.xlsx'
wb=load_workbook(path)
tabs=recuperate(path)
ws2 = wb['Calculus']
pasteRange1(1,6,ws2,tabs)
print('done',len(tabs),len(tabs[0])) # 169    14 
wb.save(path)
