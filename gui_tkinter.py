from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import glob
from tkinter import messagebox
root=Tk()
root.title('Excel Management Interface')
root.geometry('500x500')
logo=Canvas(root, bg="#4688E1")
logo.place(x=60,y=0)
wb=Workbook()
wb=load_workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/test_tk.xlsx')
ws=wb.active
a=ws['A']
b=ws['B']
c=ws['C']
dic={}
var=StringVar()
#https://blog.aspose.com/fr/create-charts-in-excel-using-python/
    # feuille = wb['Sheet1]
    # val= feuille ['F5'].value
    # vals.append(val) add element to dict
    #range(cells(1,1),cells(10,10))
    #>>> d = ws.cell(row=4, column=2, value=10)
     #for row in ws.values:
#       for value in row:
#       print(value)
    #for x in range(1,101):
#        for y in range(1,101):
#            ws.cell(row=x, column=y)
# print(Cell(1,1).value)

## add a simple formula
# ws["A1"] = "=SUM(1, 1)"

# ws.merge_cells('A2:D2')
# ws.unmerge_cells('A2:D2')
# 

def msg():
    messagebox.showinfo("Info", " Hola ! come estas ! ")
    
def get_value():
    results ='' 
    e_text=entry.get()
    cmp=0
    for cell in a:
        if cell.value==e_text:
            print(e_text)
            print(cell.value)
            results=f'{results + str(cell.value)}'
            for i in a:
                if cell.value==e_text:
                    cmp=cmp+1                
            results=f'{str(cell.value) + str(cmp) }'
    label(root, text=results)

def delete_entry():
        entry.delete(0,END)
        label.delete(0,END)

def changeText():
    label.configure(text="t'as changer le text label")
def read_label():
    print(label.cget("text"))


entry= ttk.Entry(root,font=('Century 12'),width=40)
entry.pack(pady= 30)

button= ttk.Button(root, text="Enter", command= get_value)
button.pack()

button2= ttk.Button(root, text="Delete", command= delete_entry)
button2.pack()

button3= ttk.Button(root, text="Message", command= msg)
button3.pack()

label=ttk.Label(root,font= ('Century 15 bold'))
label.pack(pady=20)

root.mainloop()