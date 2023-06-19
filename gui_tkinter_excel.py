from tkinter import *
from tkinter import filedialog
# from tkinter import tkFileDialog
# from tkFileDialog import askopenfilename 
from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter

root=Tk()
root.title('Excel Management Interface')
root.geometry('500x500')

wb=Workbook()
wb=load_workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/test_tk.xlsx')
ws=wb.active
a=ws['A']
b=ws['B']
c=ws['C']
var=StringVar()

def get_a():
    list=''
    for cell in a:
        list=f'{list + str(cell.value)} \n'
    label_a.config(text=list)

e=Entry(root, textvariable=var) 
e.pack(pady=20)

m=Message(root, text='Here you can search for your status Tests')
    
def search(): 
    results ='' 
    for cell in a:
        if cell.value=='aicha' :
            results=f'{results + str(cell.value)}\n'
        else:
            results='nothing'
    label_b.config(text=results)

label_b=Label(root, text='')
label_b.pack(pady=20)

ba=Button(root, text='Get that Column A',command=get_a)
ba.pack(pady=20)

bvar=Button(root, text='Get that text',command=search())
bvar.pack(pady=20)

label_a=Label(root, text='')
label_a.pack(pady=20)

# ws['A8']= 'Ghazal'
# ws['B8']=  24
# ws['C8']=  'move'

wb.save('test_tk1.xlsx')
root.mainloop()
