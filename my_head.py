from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate

def my_head(path,DataRoom):
    wb=load_workbook(path) #,read_only=False, keep_vba=True)
    ws = wb[DataRoom]
    maxr=ws.max_row
    maxc=ws.max_column
    info=infosheet(maxr,maxc,ws)
    endCol=info[5]
    endCol[-1]=endCol[-1]-1
    head=copyRange(1,5,endCol[0],5,ws)    # juste une seule fois
    return head
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# print(my_head(path,'DataRoom'))
