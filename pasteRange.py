from copyRange import copyRange
from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
def pasteRange(startCol, startRow, endCol, endRow, ws2,rang):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            ws2.cell(row = i, column = j).value = rang[countRow][countCol]
            countCol += 1
        countRow += 1


# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template1.xlsx'
# wb=load_workbook(path)
# ws = wb['DataRoom1']
# ws2 = wb['Calculus1']
# rang=copyRange(1,6,14,174,ws)
# pasteRange(1,6,14,174,ws2,rang)
# wb.save(path)
# print('done',len(rang),len(rang[0])) # 169    14 
