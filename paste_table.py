from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange1 import pasteRange1
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate
from Tab_Sheets import Tab_sheets
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from copy_head import  copy_head

#**************************pyuic6 -x CopyData.ui -o CopyData.py*****************
def paste_table(path,ws2,tabs):
    wb=load_workbook(path) # ,read_only=False, keep_vba=True)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1,end_column=14)
    ws2.merge_cells('A4:D4')  
    c1=ws2.cell(row=1,column=1)
    c4=ws2.cell(row=4,column=1)
    ws2['A1']= '  onsemi Internal Use Only'
    ws2['A4']= '   Test Conditions    '
    c1.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    c4.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
      

    pasteRange1(1,6,ws2,tabs)
    FullRange = "A6:" + get_column_letter(ws2.max_column) + str(ws2.max_row)
    ws2.auto_filter.ref=FullRange
   
    #wb.save(path)
      
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template1.xlsx'
# wb=load_workbook(path)
# tabs=copy_head(path,'DataRoom') 
# ws2=wb['Calculus']
# paste_table(path,ws2,tabs)
# wb.save(path)

