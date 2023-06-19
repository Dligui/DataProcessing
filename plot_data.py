from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import glob
from tkinter import messagebox
# Créer un objet Workbook
workbook = Workbook('C:/Users/zbptmj/Desktop/newdesk/py_gui/test_tk.xlsx')

# Obtenir la référence de la première feuille de calcul et ajouter des données
worksheet = workbook.getWorksheets().get(0)
worksheet.getCells().get("A2").putValue("Category1")
worksheet.getCells().get("A3").putValue("Category2")
worksheet.getCells().get("A4").putValue("Category3")
worksheet.getCells().get("B1").putValue("Column1")
worksheet.getCells().get("B2").putValue(300)
worksheet.getCells().get("B3").putValue(400)
worksheet.getCells().get("B4").putValue(200) 
worksheet.getCells().get("C1").putValue("Column2")
worksheet.getCells().get("C2").putValue(180)
worksheet.getCells().get("C3").putValue(240)
worksheet.getCells().get("C4").putValue(450)

# Ajouter un histogramme à la feuille de calcul
chartIndex = worksheet.getCharts().add(ChartType.COLUMN, 6, 2, 22, 10)

# Accéder à l'instance du graphique nouvellement ajouté
chart = worksheet.getCharts().get(chartIndex)

# Définir la source de données du graphique sur la plage "A1: B4"
chart.setChartDataRange("A1:C4", True)

# Enregistrez le fichier Excel
workbook.save("excel-column-chart.xlsx")
