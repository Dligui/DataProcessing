import xlrd
import matplotlib.pyplot as plt

document = xlrd.open_workbook("C:/Users/zbptmj/Desktop/newdesk/py_gui/temp_py.xlsx")

print("Nombre de feuilles: "+str(document.nsheets))
print("Noms des feuilles: "+str(document.sheet_names()))

feuille_1 = document.sheet_by_index(0)
feuille_1 = document.sheet_by_name("Fonction 1")

print("Format de la feuille 1:")
print("Nom: "+str(feuille_1.name))
print("Nombre de lignes: "+str(feuille_1.nrows))
print("Nombre de colonnes: "+str(feuille_1.ncols))

cols = feuille_1.ncols
rows = feuille_1.nrows

X = []
Y= []

for r in range(1, rows):
    X += [feuille_1.cell_value(rowx=r, colx=0)]
    Y += [feuille_1.cell_value(rowx=r, colx=1)]

plt.plot(X, Y)
plt.show()


import openpyxl

#Prepare the spreadsheets to copy from and paste too.

#File to be copied
wb = openpyxl.load_workbook("foo.xlsx") #Add file name
sheet = wb.get_sheet_by_name("foo") #Add Sheet name

#File to be pasted into
template = openpyxl.load_workbook("foo2.xlsx") #Add file name
temp_sheet = template.get_sheet_by_name("foo2") #Add Sheet name
#******************************************************need this*********************************************
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):

            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
def createData():
    print("Processing...")
    selectedRange = copyRange(1,2,4,14,sheet) #Change the 4 number values
    pastingRange = pasteRange(1,3,4,15,temp_sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("foo.xlsx")
    print("Range copied and pasted!")`