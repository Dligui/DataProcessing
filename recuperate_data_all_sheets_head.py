from openpyxl.workbook import Workbook
from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
import xlrd
from collections import defaultdict
from copyall import copyall
from copyRange import copyRange
from pasteRange import pasteRange
from getTable import getTable
from infosheet import infosheet
from rangindex import rangindex
from openpyxl.styles import Alignment
from recuperate_data_all_sheets import recuperate

#**************************pyuic6 -x CopyData.ui -o CopyData.py*****************
def copy_head(path,DataRoom):
    wb=load_workbook(path) # ,read_only=False, keep_vba=True)
    ws = wb[DataRoom]
    maxr=ws.max_row
    maxc=ws.max_column

    info=infosheet(maxr,maxc,ws)
    endCol=info[5]
    endCol[-1]=endCol[-1]-1
    tab=[]
    head=copyRange(1,5,endCol[0],5,ws)    # juste une seule fois
    tab=recuperate(path)
    head.extend(tab)
    return head


#function test
# path='C:/Users/zbptmj/Desktop/newdesk/py_gui/new_template.xlsx'
# wb=load_workbook(path)
# ws = 'DataTemp'
# tab=copy_head(path,ws)
# print(tab)
